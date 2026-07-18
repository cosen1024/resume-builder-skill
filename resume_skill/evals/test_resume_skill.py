import importlib.util
import io
import tempfile
import unittest
from contextlib import redirect_stderr
from pathlib import Path
from unittest import mock

import frontmatter


SKILL_DIR = Path(__file__).resolve().parents[1]


def load_script(name):
    path = SKILL_DIR / "scripts" / f"{name}.py"
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


csv_to_md = load_script("csv_to_md")
render = load_script("render")


class CsvToMarkdownTests(unittest.TestCase):
    def test_front_matter_handles_colons_newlines_and_numeric_strings(self):
        meta = {
            "name": "张三: 后端",
            "phone": "01234567890",
            "email": "test@example.com",
        }
        md = csv_to_md.to_markdown(
            meta,
            {"期望职位": "研发: 后端"},
            "第一行\n第二行",
            [],
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.md"
            path.write_text(md, encoding="utf-8")
            post = frontmatter.load(path)

        self.assertEqual(post["name"], "张三: 后端")
        self.assertEqual(post["phone"], "01234567890")
        self.assertEqual(post["intent"], "研发: 后端")
        self.assertEqual(post["summary"], "第一行\n第二行")

    def test_legacy_xls_is_rejected_with_clear_message(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.xls"
            path.write_bytes(b"not-an-xlsx-file")
            with self.assertRaisesRegex(SystemExit, "不支持旧版 .xls"):
                csv_to_md.read_rows(path)

    def test_bundled_csv_example_builds_a_renderable_resume(self):
        rows = csv_to_md.read_rows(SKILL_DIR / "assets" / "resume.example.csv")
        meta, intent_parts, summary, groups = csv_to_md.build(rows)
        md = csv_to_md.to_markdown(meta, intent_parts, summary, groups)

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.md"
            path.write_text(md, encoding="utf-8")
            parsed_meta, sections = render.parse_resume(path)

        render.validate_resume(parsed_meta, sections)
        self.assertEqual(parsed_meta["name"], "库森")
        self.assertGreaterEqual(len(sections), 4)

    def test_xlsx_input_is_supported(self):
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["分类", "字段名", "值"])
            sheet.append(["基本信息", "姓名", "库森"])
            workbook.save(path)

            rows = csv_to_md.read_rows(path)

        self.assertEqual(rows[1], ["基本信息", "姓名", "库森"])

    def test_xlsx_rejects_numeric_values_beyond_excel_precision(self):
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["分类", "字段名", "值"])
            sheet.append(["基本信息", "身份证号", 330100199901010000])
            workbook.save(path)

            with self.assertRaisesRegex(SystemExit, "设为“文本”"):
                csv_to_md.read_rows(path)

    def test_private_id_is_commented_out(self):
        md = csv_to_md.to_markdown(
            {"name": "库森", "id_number": "330100199901010000"},
            {},
            "",
            [],
        )

        post = frontmatter.loads(md)
        self.assertNotIn("id_number", post.metadata)
        self.assertIn("# id_number:", md)

    def test_unknown_category_digit_suffix_is_stripped_from_entry_heading(self):
        rows = [
            ["分类", "字段名", "值"],
            ["基本信息", "姓名", "库森"],
            ["竞赛经历1", "奖项名称", "全国大学生数学建模竞赛一等奖"],
        ]
        meta, intent_parts, summary, groups = csv_to_md.build(rows)
        md = csv_to_md.to_markdown(meta, intent_parts, summary, groups)

        self.assertNotIn("竞赛经历1", md)
        self.assertIn("## 竞赛经历", md)
        self.assertIn("### 竞赛经历", md)
        self.assertIn("- 全国大学生数学建模竞赛一等奖", md)

    def test_xlsx_datetime_cells_keep_only_the_date(self):
        import datetime
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "resume.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["分类", "字段名", "值"])
            sheet.append(["基本信息", "出生日期", datetime.datetime(1999, 11, 1)])
            workbook.save(path)

            rows = csv_to_md.read_rows(path)

        self.assertEqual(rows[1], ["基本信息", "出生日期", "1999-11-01"])
        self.assertNotIn("00:00:00", rows[1][2])
    def test_meta_keys_map_contact_link_fields_case_insensitively(self):
        meta, _, _, _ = csv_to_md.build([
            ["基本信息", "微信", "kusen-wx"],
            ["基本信息", "个人主页", "https://example.com"],
            ["基本信息", "GitHub", "gh-user"],
            ["基本信息", "领英", "li-user"],
        ])
        self.assertEqual(meta["wechat"], "kusen-wx")
        self.assertEqual(meta["website"], "https://example.com")
        self.assertEqual(meta["github"], "gh-user")
        self.assertEqual(meta["linkedin"], "li-user")

        # 同义字段名与大小写变体映射到同一个 key
        for field in ("微信号", "wechat", "WECHAT", "WeChat"):
            meta, _, _, _ = csv_to_md.build([["基本信息", field, "v"]])
            self.assertEqual(meta.get("wechat"), "v", field)
        for field in ("github", "GITHUB"):
            meta, _, _, _ = csv_to_md.build([["基本信息", field, "v"]])
            self.assertEqual(meta.get("github"), "v", field)
        for field in ("linkedin", "LinkedIn"):
            meta, _, _, _ = csv_to_md.build([["基本信息", field, "v"]])
            self.assertEqual(meta.get("linkedin"), "v", field)
        for field in ("主页", "博客", "个人网站", "website", "Website"):
            meta, _, _, _ = csv_to_md.build([["基本信息", field, "v"]])
            self.assertEqual(meta.get("website"), "v", field)

        # 生成的 Markdown 能把新字段写进 front matter 并被解析回来
        md = csv_to_md.to_markdown(
            {"name": "库森", "wechat": "kusen-wx", "github": "gh-user",
             "website": "https://example.com", "linkedin": "li-user"},
            {},
            "",
            [],
        )
        post = frontmatter.loads(md)
        self.assertEqual(post["wechat"], "kusen-wx")
        self.assertEqual(post["github"], "gh-user")
        self.assertEqual(post["website"], "https://example.com")
        self.assertEqual(post["linkedin"], "li-user")


class RenderTests(unittest.TestCase):
    def test_template_discovery_only_returns_renderable_templates(self):
        self.assertEqual(
            render.available_templates(),
            ["classic", "compact", "minimal", "modern", "timeline"],
        )

    def test_html_escapes_metadata_titles_and_headings(self):
        meta = {"name": "<测试 & 用户>", "intent": "研发 <后端>"}
        sections = [{
            "title": "项目 <经验>",
            "bullets": [],
            "entries": [{
                "title_parts": ["A & B", "<核心开发>"],
                "meta": "2024 < 2025",
                "bullets": [render.inline_md("使用 **A&B** 与 `<unsafe>`")],
            }],
        }]

        output = render.render_html(meta, sections, "classic")

        self.assertIn("&lt;测试 &amp; 用户&gt;", output)
        self.assertIn("项目 &lt;经验&gt;", output)
        self.assertIn("<strong>A &amp; B</strong>", output)
        self.assertIn("&lt;核心开发&gt;", output)
        self.assertIn("2024 &lt; 2025", output)
        self.assertIn("<strong>A&amp;B</strong>", output)
        self.assertIn("<code>&lt;unsafe&gt;</code>", output)

    def test_inline_markdown_has_valid_triple_emphasis_order(self):
        output = str(render.inline_md("***粗斜体***"))
        self.assertEqual(output, "<strong><em>粗斜体</em></strong>")

    def test_inline_markdown_link_renders_anchor(self):
        output = str(render.inline_md("[博客](https://example.com)"))
        self.assertEqual(output, '<a href="https://example.com">博客</a>')

    def test_inline_markdown_link_supports_mailto(self):
        output = str(render.inline_md("[邮箱](mailto:a@b.com)"))
        self.assertEqual(output, '<a href="mailto:a@b.com">邮箱</a>')

    def test_inline_markdown_link_scheme_whitelist_is_case_insensitive(self):
        self.assertIn('href="HTTPS://example.com"',
                      str(render.inline_md("[x](HTTPS://example.com)")))
        self.assertIn('href="Mailto:a@b.com"',
                      str(render.inline_md("[x](Mailto:a@b.com)")))

    def test_inline_markdown_link_rejects_disallowed_schemes(self):
        for url in ("javascript:alert(1)", "data:text/html,<b>x</b>",
                    "file:///etc/passwd", "vbscript:msgbox(1)"):
            with self.subTest(url=url):
                output = str(render.inline_md(f"[点我]({url})"))
                self.assertNotIn("href", output)
                self.assertNotIn("<a", output)
                self.assertIn("点我", output)  # 退化为纯文字

    def test_inline_markdown_link_escapes_label_and_url(self):
        output = str(render.inline_md('[a<b & "c"](https://example.com/?a=1&b="2")'))
        self.assertIn('<a href="https://example.com/?a=1&amp;b=&quot;2&quot;">', output)
        self.assertIn('a&lt;b &amp; &quot;c&quot;</a>', output)

    def test_inline_markdown_link_url_special_chars_survive_emphasis(self):
        # URL 里的 *、` 不应被强调/代码语法匹配破坏
        output = str(render.inline_md("[x](https://example.com/*star*`tick`)"))
        self.assertIn('href="https://example.com/*star*`tick`"', output)
        self.assertNotIn("<em>", output)
        # 链接外的强调语法仍正常工作
        output = str(render.inline_md("**重点**见[博客](https://example.com)"))
        self.assertEqual(
            output,
            '<strong>重点</strong>见<a href="https://example.com">博客</a>',
        )

    def test_inline_markdown_link_relative_url_is_rejected(self):
        output = str(render.inline_md("[x](/local/path)"))
        self.assertNotIn("href", output)
        self.assertIn("x", output)

    def test_heading_accepts_ascii_or_fullwidth_separators_without_spaces(self):
        ascii_heading = render.parse_entry_heading("### 公司|2024.01-2024.06")
        fullwidth_heading = render.parse_entry_heading("### 公司·部门｜2024.01-2024.06")
        alternate_heading = render.parse_entry_heading("### 公司・部门•岗位 | 2025.01-2025.06")

        self.assertEqual(ascii_heading, {
            "title_parts": ["公司"],
            "meta": "2024.01-2024.06",
        })
        self.assertEqual(fullwidth_heading, {
            "title_parts": ["公司", "部门"],
            "meta": "2024.01-2024.06",
        })
        self.assertEqual(alternate_heading, {
            "title_parts": ["公司", "部门", "岗位"],
            "meta": "2025.01-2025.06",
        })

    def test_invalid_accent_is_rejected(self):
        with self.assertRaises(SystemExit):
            render.render_html({}, [], "modern", accent="red; } body { color: red")

    def test_orange_accent_preset_is_injected_into_template_css(self):
        self.assertEqual(render.ACCENT_PRESETS.get("orange"), "#c65d21")

        output = render.render_html({}, [], "modern", accent="orange")

        self.assertIn(":root { --accent: #c65d21; }", output)

    def test_custom_hex_accent_is_supported(self):
        output = render.render_html({}, [], "modern", accent="#1f6feb")
        self.assertIn(":root { --accent: #1f6feb; }", output)

    def test_classic_template_remains_accent_independent(self):
        classic_css = (
            SKILL_DIR / "assets" / "templates" / "classic" / "style.css"
        ).read_text(encoding="utf-8")
        compact_css = (
            SKILL_DIR / "assets" / "templates" / "compact" / "style.css"
        ).read_text(encoding="utf-8")

        self.assertNotIn("var(--accent)", classic_css)
        self.assertIn("var(--accent)", compact_css)
        self.assertIn("font-family: var(--font-serif)", classic_css)
        self.assertIn(".sec-title {\n  font-family: var(--font-sans)", classic_css)

    def test_optional_photo_is_rendered_by_supported_templates(self):
        for template in ("classic", "compact", "modern", "minimal"):
            with self.subTest(template=template):
                output = render.render_html({"name": "库森", "photo": "avatar.png"}, [], template)
                self.assertIn('src="avatar.png"', output)

        classic = render.render_html({"name": "库森", "photo": "avatar.png"}, [], "classic")
        self.assertIn('class="hd has-photo"', classic)

        timeline = render.render_html({"name": "库森", "photo": "avatar.png"}, [], "timeline")
        self.assertNotIn('src="avatar.png"', timeline)

    def test_intent_detail_is_rendered_by_all_templates(self):
        meta = {
            "name": "库森",
            "intent": "后端开发工程师",
            "intent_detail": "期望地点：杭州 · 到岗时间：2026.06",
        }
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html(meta, [], template)
                self.assertIn("期望地点：杭州", output)

    def test_hometown_is_rendered_by_all_templates(self):
        meta = {"name": "库森", "hometown": "浙江省"}
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html(meta, [], template)
                self.assertIn("户籍：浙江省", output)

    def test_hometown_leaves_no_placeholder_when_absent(self):
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html({"name": "库森"}, [], template)
                self.assertNotIn("hometown", output)

    def test_contact_link_fields_render_in_all_templates(self):
        meta = {
            "name": "库森",
            "wechat": "wx-kusen",
            "github": "gh-kusen",
            "website": "https://blog.kusen.dev",
            "linkedin": "li-kusen",
        }
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html(meta, [], template)
                for value in ("wx-kusen", "gh-kusen",
                              "https://blog.kusen.dev", "li-kusen"):
                    self.assertIn(value, output)

    def test_contact_link_fields_are_omitted_when_absent(self):
        # 字段缺省时不输出值，也不残留字段名/标签等标记
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html({"name": "库森"}, [], template)
                for marker in ("wechat", "github", "website", "linkedin"):
                    self.assertNotIn(marker, output)

    def test_contact_link_fields_render_partial_subset(self):
        # 只给部分字段时，仅输出给了的字段
        meta = {"name": "库森", "github": "gh-kusen"}
        for template in render.available_templates():
            with self.subTest(template=template):
                output = render.render_html(meta, [], template)
                self.assertIn("gh-kusen", output)
                self.assertNotIn("wechat", output)
                self.assertNotIn("linkedin", output)
>>>>>>> 7a65544 (feat: add inline links, social contact fields, PDF page count)

    def test_resume_requires_a_name(self):
        stderr = io.StringIO()
        with redirect_stderr(stderr), self.assertRaises(SystemExit):
            render.validate_resume({}, [])
        self.assertIn("缺少必填字段 name", stderr.getvalue())

    def test_old_python_version_is_rejected_clearly(self):
        stderr = io.StringIO()
        with redirect_stderr(stderr), self.assertRaises(SystemExit):
            render.require_python_version((3, 9, 18))
        self.assertIn("需要 Python 3.10 或更高版本", stderr.getvalue())

    def test_weasyprint_system_dependency_error_is_not_reported_as_missing_package(self):
        real_import = __import__

        def import_with_weasyprint_error(name, *args, **kwargs):
            if name == "weasyprint":
                raise OSError("libgobject-2.0 not found")
            return real_import(name, *args, **kwargs)

        stderr = io.StringIO()
        with mock.patch("builtins.__import__", side_effect=import_with_weasyprint_error):
            with redirect_stderr(stderr), self.assertRaises(SystemExit):
                render.require_deps(need_pdf=True)

        output = stderr.getvalue()
        self.assertIn("系统图形库加载失败", output)
        self.assertIn("--html-only", output)
        self.assertNotIn("缺少依赖", output)

    def test_unknown_front_matter_field_warns_with_alias(self):
        stderr = io.StringIO()
        with redirect_stderr(stderr):
            render.validate_resume(
                {"name": "库森", "求职意向": "后端开发工程师"},
                [],
            )
        output = stderr.getvalue()
        self.assertIn("字段 `求职意向` 不会被模板使用", output)
        self.assertIn("是否想写 `intent`", output)
        self.assertIn("没有求职意向", output)

    def test_missing_intent_warns_but_does_not_fail(self):
        stderr = io.StringIO()
        with redirect_stderr(stderr):
            render.validate_resume({"name": "库森"}, [])
        self.assertIn("没有求职意向", stderr.getvalue())

    def test_template_specific_html_output_avoids_overwrite(self):
        md_path = Path("resume.md")
        self.assertEqual(
            render.html_output_path(md_path, "classic"),
            Path("resume.classic.html"),
        )
        self.assertEqual(
            render.html_output_path(md_path, "modern"),
            Path("resume.modern.html"),
        )

    def test_ensure_parent_creates_output_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "nested" / "resume.pdf"
            render.ensure_parent(output)
            self.assertTrue(output.parent.is_dir())

    def test_bundled_markdown_example_is_one_a4_page_in_all_templates(self):
        from weasyprint import HTML

        md_path = SKILL_DIR / "assets" / "resume.example.md"
        meta, sections = render.parse_resume(md_path)
        render.validate_resume(meta, sections)

        for template in render.available_templates():
            with self.subTest(template=template):
                html = render.render_html(meta, sections, template)
                document = HTML(string=html, base_url=str(md_path.parent)).render()
                self.assertEqual(len(document.pages), 1)
                self.assertAlmostEqual(document.pages[0].width, 793.7, delta=1)
                self.assertAlmostEqual(document.pages[0].height, 1122.5, delta=1)

    def test_long_resume_paginates_to_multiple_a4_pages(self):
        from weasyprint import HTML

        md_path = SKILL_DIR / "assets" / "resume.example.md"
        meta, sections = render.parse_resume(md_path)
        long_sections = sections * 3

        for template in render.available_templates():
            with self.subTest(template=template):
                html = render.render_html(meta, long_sections, template)
                document = HTML(string=html, base_url=str(md_path.parent)).render()
                self.assertGreaterEqual(len(document.pages), 2)
                for page in document.pages:
                    self.assertAlmostEqual(page.width, 793.7, delta=1)
                    self.assertAlmostEqual(page.height, 1122.5, delta=1)

    def test_html_to_pdf_reports_one_page_for_example_resume(self):
        md_path = SKILL_DIR / "assets" / "resume.example.md"
        meta, sections = render.parse_resume(md_path)
        render.validate_resume(meta, sections)
        html = render.render_html(meta, sections, "classic")

        with tempfile.TemporaryDirectory() as tmp:
            pdf_path = Path(tmp) / "resume.pdf"
            pages = render.html_to_pdf(html, pdf_path, base_url=md_path.parent)
            self.assertEqual(pages, 1)
            self.assertTrue(pdf_path.is_file())
            self.assertGreater(pdf_path.stat().st_size, 0)

    def test_html_to_pdf_reports_multiple_pages_for_long_resume(self):
        md_path = SKILL_DIR / "assets" / "resume.example.md"
        meta, sections = render.parse_resume(md_path)
        html = render.render_html(meta, sections * 3, "classic")

        with tempfile.TemporaryDirectory() as tmp:
            pdf_path = Path(tmp) / "resume.pdf"
            pages = render.html_to_pdf(html, pdf_path, base_url=md_path.parent)
            self.assertGreaterEqual(pages, 2)
            self.assertTrue(pdf_path.is_file())


if __name__ == "__main__":
    unittest.main()
